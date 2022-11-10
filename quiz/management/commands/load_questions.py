from openpyxl import load_workbook
import datetime
import os.path
from django.core.management.base import BaseCommand, CommandError
from quiz.models import Question, Answer 

def load_partition_questions(id_task, lang, partition_number, subpartition_number, file_name):
        theme_number = 2
        s_now = datetime.datetime.now()
        partition_messages = []
        # if appcfg.os == 'unix':
        #    file_path = appcfg.UPLOAD_PATH + '/' + lang + '/' + file_name
        # else:
        #    file_path = appcfg.UPLOAD_PATH + '\\' + lang + '\\' + file_name

        #
        # if appcfg.os == 'unix':
        file_path = "C:/Users/Admin/Desktop/pdd" + '/' + lang +  '/' + file_name
        # file_path = "C:\\Users\\Admin\\Desktop\\pdd" + '\\' + lang +  '\\' + file_name
        path = os.path.normpath(file_path)

        # print("Download started: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)


        if not os.path.isfile(file_path):
            l_mess = f"ERROR ! File not exists: {file_path}"
            partition_messages.append(l_mess)
            # print(l_mess)
            # print("---")
            # return partition_messages

        wb = load_workbook(path)
        l_mess = f"Загружен Excel file: {file_path}"
        # print(l_mess)
        # partition_messages.append(l_mess)

        sheet = wb.active
        # cursor.callproc("pdd_testing.load_questions.clean", [id_task, theme_number, partition_number, subpartition_number])
        l_mess = f"Очистили вопросы: id_task: {id_task}, theme_number: {theme_number}, " \
                f"partition_number: {partition_number}, subpartition_number: {subpartition_number}"
        # print(l_mess)
        # partition_messages.append(l_mess)

        # Создадим новое задание
        # file_split = os.path.splitext(file_name)
        # id_theme = cursor.callfunc('admin.theme_new', int, (id_task, file_split[0]))
        # if not id_theme:
        #     print('Ошибка регистрации нового задания...')
        id_quest = 0
        id_prev_quest = -1
        order_num = 0
        id_question = 0
        for i in range(2, sheet.max_row+1):
            id_curr_quest = sheet.cell(row=i, column=1).value
            quest = sheet.cell(row=i, column=2).value
            correctly = sheet.cell(row=i, column=3).value
            answer = sheet.cell(row=i, column=4).value
            url_image = sheet.cell(row=i, column=5).value
            order_num = order_num + 1

            if not id_curr_quest:
                l_mess = f"{file_name}: WARNING ! FINISH. Преждевременное завершение загрузки. " \
                        f"Последняя загруженная строка: {id_prev_quest}"
                # partition_messages.append(l_mess)
                # print(l_mess)
                break
            # if id_curr_quest != id_prev_quest:
            try:
                if quest:
                    id_quest = id_quest + 1
                    order_num = 1
                    question = Question.objects.create(id_task=id_task,
                    theme_number=theme_number,  partition_number=partition_number, subpartition_number=subpartition_number,
                    url_image=url_image, order_num_question=order_num, question=quest)
                    question.save()
                    id_question = Question.objects.filter(id=id_quest)
                    print("pdd_testing.load_questions.add_question", 
                                                [id_task, theme_number,
                                                partition_number, subpartition_number, id_quest, url_image, quest])
                if int(id_question) > 0:
                    upload_answer = Answer.objects.create(id_question=id_question, order_num=order_num, 
                    correctly=correctly, answer=answer)
                    upload_answer.save()
                    print("pdd_testing.load_questions.add_answer", [id_question, order_num, correctly, answer])
                id_prev_quest = id_curr_quest
            except BaseException as e:
                l_mess = f"{file_name}: ERROR ! Empty id_question, Номер вопроса: {order_num}, error: {e}"
                partition_messages.append(l_mess)
                print(l_mess)
                break

        # con.commit()
        # con.close()
        now = datetime.datetime.now()
        l_mess = f"Загрузка вопросов завершена: {file_name}. " \
                f"Последняя загруженная строка: {id_prev_quest} : {now.strftime('%d-%m-%Y %H:%M:%S')}"
        # print(l_mess)
        partition_messages.append(l_mess)
        return partition_messages

class Command(BaseCommand):
    def load_task(id_task, lang):
        part = 1
        load_messages = []
        for sub_part in range(1, 15):
            f_name = str(sub_part) + ' подраздел.xlsx'
            load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'ОБД.xlsx'
        part = 2
        load_messages.append(load_partition_questions(id_task, lang, part, 0, f_name))

        # f_name = 'Медицина.xlsx'
        # part = 3
        # load_messages.append(load_partition_questions(id_task, lang, part, 0, f_name))

        f_name = 'Административная ответственность.xlsx'
        part = 4
        load_messages.append(load_partition_questions(id_task, lang, part, 0, f_name))

        f_name = 'ПДДАА1В1.xlsx'
        part = 5
        sub_part = 1
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'ПДДВВЕ.xlsx'
        part = 5
        sub_part = 2
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'ПДДС1С.xlsx'
        part = 5
        sub_part = 3
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'ПДДD1DТb.xlsx'
        part = 5
        sub_part = 4
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'ПДДC1ECED1EDE.xlsx'
        part = 5
        sub_part = 5
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'ПДД Tm.xlsx'
        part = 5
        sub_part = 6
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'СПОБДА1АВ1.xlsx'
        part = 6
        sub_part = 1
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'СПОБДС1С.xlsx'
        part = 6
        sub_part = 3
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'СПОБДD1DTb.xlsx'
        part = 6
        sub_part = 4
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'СПОБДC1ECED1EDE.xlsx'
        part = 6
        sub_part = 5
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))

        f_name = 'СПОБДTm.xlsx'
        part = 6
        sub_part = 6
        load_messages.append(load_partition_questions(id_task, lang, part, sub_part, f_name))
        return load_messages


    load_task(1, 'ru')